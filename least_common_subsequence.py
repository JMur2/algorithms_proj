import pandas as pd

from openpyxl import load_workbook 
from Player import Player

class Project:

    def __init__(self):
        # data frames for storing player data
        self.pre_move_df = None
        self.post_move_df = None
        self.players = []
        self.pre_weights = None
        self.post_weights = None
        self.baseline = None

    def read_data(self):
        # makes use of the openpyxl library to read in data from excel
        wb = load_workbook(filename='data.xlsx', read_only=True)
        ws = wb['Sheet1']

        # selects the stats of players before they moved teams
        data_rows = []
        for row in ws['C4':'I63']:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)
        
        self.pre_move_df = pd.DataFrame(data_rows)

        # selects the stats of players after they moved teams
        data_rows = []
        for row in ws['K4':'Q63']:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)
        
        self.post_move_df = pd.DataFrame(data_rows)

        data_rows = []
        for row in ws['C93':'K97']:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)
        
        self.pre_weights = pd.DataFrame(data_rows)

        data_rows = []
        for row in ws['W93':'AE97']:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)
        
        self.post_weights = pd.DataFrame(data_rows)

        data_rows = []
        for row in ws['O93':'S93']:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)
        
        self.baseline = pd.DataFrame(data_rows)
        
    def set_all_player_stats(self):
        """
        This function looks at the set of players found in the excel document, and creates a player object for each unique player.
        The player object stores a player id (their number in the list), as well as the players 2 sets of stats
        """
        for i in range(60):
            p = Player(i)

            p.set_pre_move_stats(self.pre_move_df.values[i])
            p.set_post_move_stats(self.post_move_df.values[i])

            self.players.append(p)

    def create_baseline(self, id):
        for p in self.players:

            if id == 0:
                arr = p.pre_stats
                weights = self.pre_weights
                print(weights)
            elif id == 1:
                arr = p.post_stats
                weights = self.post_weights

            age = arr[0]
            pos = arr[1]
            
            holder = []
            if age <= 24:
                for i in range(5):
                    holder.append(self.baseline[i][0] * weights[6][i])
            elif age >= 25 and age <= 28:
                for i in range(5):
                    holder.append(self.baseline[i][0] * weights[7][i])
            else:
                for i in range(5):
                    holder.append(self.baseline[i][0] * weights[8][i])

            if pos == 'PG':
                for i in range(5):
                    holder[i] = holder[i] * weights[0][i]
            elif pos == 'SG':
                for i in range(5):
                    holder[i] = holder[i] * weights[1][i]
            elif pos == 'SF':
                for i in range(5):
                    holder[i] = holder[i] * weights[2][i]
            elif pos == 'PF':
                for i in range(5):
                    holder[i] = holder[i] * weights[3][i]
            elif pos == 'C':
                for i in range(5):
                    holder[i] = holder[i] * weights[4][i]
            
            score = []
            for i in range(5):
                if holder[i] <= arr[i+2]:
                    score.append(1)
                else:
                    score.append(0)
            
            score = pd.DataFrame(score)
            if id == 0:
                p.set_pre_stat_scores(score[0])
            elif id == 1:
                p.set_post_stat_scores(score[0])
            
    def lcs(self, X, Y, m, n):
        """
        This algorithm checks for the longest common subsequence between 2 arrays and includes a modification that places a higher 
        weights on what will be a "good stat" on a player

        Solved with recursion

        Parameters:
        X: array of values
        Y: array of values
        """
        if m == 0 or n == 0:
            return 0
        if X[m-1] == 0:
            return 1 + self.lcs(X, Y, m-1, n-1)
        else:
            return max(self.lcs(X, Y, m, n-1), self.lcs(X, Y, m-1, n))

if __name__ == "__main__":

    for i in range(5):
        print(i)

    project = Project()

    project.read_data()

    project.set_all_player_stats()

    # print(project.weights)
    # print(project.weights[6][1])

    # print(project.baseline)
    # print(project.baseline[2][0])

    project.create_baseline(0)
    project.create_baseline(1)

    lcs_score = []
    for p in project.players:
        lcs_score.append(project.lcs(p.pre_scores, p.post_scores, 5, 5))

    print(lcs_score)
